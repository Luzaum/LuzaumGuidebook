from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_GLOB = "*GENUTRI*.xlsx"
DEFAULT_OUTPUT = ROOT / "modules" / "energia-vet" / "data" / "genutri-dataset.json"
DEFAULT_AUDIT = ROOT / "docs" / "energia-vet-genutri-audit.md"


FOOD_HEADER_MAP = {
    "Alimentos ": "name",
    "Categoria": "category",
    "Umidade (%)": "moisturePct",
    "Matéria Seca (%)": "dryMatterPct",
    "Energia (kcal)/100g": "energyKcalPer100g",
    "Proteína Bruta (%)": "crudeProteinPct",
    "Extrato Etéreo (%)": "etherExtractPct",
    "Matéria Mineral (%)": "ashPct",
    "Fibra Bruta (%)": "crudeFiberPct",
    "Extrativo Não Nitrogenado (%)": "nitrogenFreeExtractPct",
    "Cálcio (%)": "calciumPct",
    "Fósforo (%)": "phosphorusPct",
    "Potássio (%)": "potassiumPct",
    "Sódio (%)": "sodiumPct",
    "Cloro (%)": "chloridePct",
    "Magnésio (%)": "magnesiumPct",
    "Manganês (mg)": "manganeseMg",
    "Cobre (mg)": "copperMg",
    "Zinco (mg)": "zincMg",
    "Selênio (mg)": "seleniumMg",
    "Vitamina A (UI)": "vitaminAIu",
    "Vitamina D (UI)": "vitaminDIu",
    "Vitamina E (UI)": "vitaminEIu",
    "Tiamina (mg)": "thiamineMg",
    "Riboflavina (mg)": "riboflavinMg",
    "Piridoxina (mg)": "pyridoxineMg",
    "Taurina (%)": "taurinePct",
    "Met + Cis (%)": "methionineCystinePct",
    "Omega 3 (%)": "omega3Pct",
    "Omega 6 (%)": "omega6Pct",
    "AGCM (%)": "mcfaPct",
    "Colina (mg)": "cholineMg",
    "Ferro (mg)": "ironMg",
    "Iodo (mg)": "iodineMg",
    "Niacina (mg)": "niacinMg",
    "Ácido Pantotenico (mg)": "pantothenicAcidMg",
    "Biotina (mg)": "biotinMg",
    "Ácido Fólico (mg)": "folicAcidMg",
    "Cobalamina (mcg)": "cobalaminMcg",
    "Vitamina K (mg)": "vitaminKMg",
}


REQUIREMENT_HEADER_MAP = {
    "Referências": "sourceLabel",
    "Nutrientes": "profileLabel",
    "Umidade (%)": "moisturePct",
    "Energia (kcal)/100g": "energyKcalPer100g",
    "Proteína Bruta (%)": "crudeProteinPct",
    "Arg (%)": "argininePct",
    "His (%)": "histidinePct",
    "Iso (%)": "isoleucinePct",
    "Leu (%)": "leucinePct",
    "Lis (%)": "lysinePct",
    "Met (%)": "methioninePct",
    "Met + Cis (%)": "methionineCystinePct",
    "Fen (%)": "phenylalaninePct",
    "Fen + Tir (%)": "phenylalanineTyrosinePct",
    "Tre (%)": "threoninePct",
    "Tri (%)": "tryptophanPct",
    "Val (%)": "valinePct",
    "Tau (%)": "taurinePct",
    "Extrato Etéreo (%)": "etherExtractPct",
    "Ômega 6 (%)": "omega6Pct",
    "Ara (mg)": "arachidonicAcidMg",
    "Ômega 3 (%)": "omega3Pct",
    "EPA ": "epaPct",
    "DHA": "dhaPct",
    "EPA + DHA (%)": "epaDhaPct",
    "6:3": "omega6To3Ratio",
    "Cálcio (%)": "calciumPct",
    "Fósforo (%)": "phosphorusPct",
    "Potássio (g)": "potassiumG",
    "Sódio (g)": "sodiumG",
    "Cloro (g)": "chlorideG",
    "Magnáesio (g)": "magnesiumG",
    "CU -mg": "copperMg",
    "I -mg": "iodineMg",
    "FE -mg": "ironMg",
    "MGN -mg": "manganeseMg",
    "Se (mg/Kg)": "seleniumMgPerKg",
    "Zn-(mg/Kg)": "zincMgPerKg",
    "VITAMINA": "vitaminSection",
    "Vit A (UI)": "vitaminAIu",
    "Vit. D (UI/Kg)": "vitaminDIuPerKg",
    "Vit. E (UI/Kg)": "vitaminEIuPerKg",
    "B1 -mg": "thiamineMg",
    "B2 -mg": "riboflavinMg",
    "B5 -mg": "pantothenicAcidMg",
    "B6 -mg": "pyridoxineMg",
    "B12 -ug": "cobalaminMcg",
    "B3 -mg": "niacinMg",
    "B9 -ug": "folateMcg",
    "Biotina": "biotinRaw",
    "COLINA": "cholineRaw",
    "K-mg": "vitaminKMg",
    "Vit. C (mg/Kg)": "vitaminCMgPerKg",
    "Carnitina (mg/Kg)": "carnitineMgPerKg",
    "pH Urinário Gato": "urinaryPhCat",
    "Densidade urinária": "urinaryDensity",
    "Fibra Bruta (%)": "crudeFiberPct",
    "Extrativo Não Nitrogenado (%)": "nitrogenFreeExtractPct",
    "Recomendações extras": "extraRecommendations",
}


NUTRIENT_UNITS = {
    "moisturePct": "%",
    "dryMatterPct": "%",
    "energyKcalPer100g": "kcal/100g",
    "crudeProteinPct": "%",
    "etherExtractPct": "%",
    "ashPct": "%",
    "crudeFiberPct": "%",
    "nitrogenFreeExtractPct": "%",
    "calciumPct": "%",
    "phosphorusPct": "%",
    "potassiumPct": "%",
    "sodiumPct": "%",
    "chloridePct": "%",
    "magnesiumPct": "%",
    "manganeseMg": "mg",
    "copperMg": "mg",
    "zincMg": "mg",
    "seleniumMg": "mg",
    "vitaminAIu": "UI",
    "vitaminDIu": "UI",
    "vitaminEIu": "UI",
    "thiamineMg": "mg",
    "riboflavinMg": "mg",
    "pyridoxineMg": "mg",
    "taurinePct": "%",
    "methionineCystinePct": "%",
    "omega3Pct": "%",
    "omega6Pct": "%",
    "mcfaPct": "%",
    "cholineMg": "mg",
    "ironMg": "mg",
    "iodineMg": "mg",
    "niacinMg": "mg",
    "pantothenicAcidMg": "mg",
    "biotinMg": "mg",
    "folicAcidMg": "mg",
    "cobalaminMcg": "mcg",
    "vitaminKMg": "mg",
    "argininePct": "%",
    "histidinePct": "%",
    "isoleucinePct": "%",
    "leucinePct": "%",
    "lysinePct": "%",
    "methioninePct": "%",
    "phenylalaninePct": "%",
    "phenylalanineTyrosinePct": "%",
    "threoninePct": "%",
    "tryptophanPct": "%",
    "valinePct": "%",
    "arachidonicAcidMg": "mg",
    "epaPct": "%",
    "dhaPct": "%",
    "epaDhaPct": "%",
    "omega6To3Ratio": "ratio",
    "potassiumG": "g",
    "sodiumG": "g",
    "chlorideG": "g",
    "magnesiumG": "g",
    "seleniumMgPerKg": "mg/kg",
    "zincMgPerKg": "mg/kg",
    "vitaminDIuPerKg": "UI/kg",
    "vitaminEIuPerKg": "UI/kg",
    "folateMcg": "mcg",
    "vitaminCMgPerKg": "mg/kg",
    "carnitineMgPerKg": "mg/kg",
    "urinaryPhCat": "pH",
    "urinaryDensity": "specific_gravity",
}


LOWERCASE_CATEGORY_FIXES = {
    "critical": "Critical",
    "sucedaneo": "Sucedâneo",
    "ração": "Ração",
}


MANUAL_FOOD_TYPE_RULES = {
    "Água": "natural",
    "Critical": "suplemento",
    "Nutra pet": "suplemento",
    "Renafil": "suplemento",
    "Glutamina": "suplemento",
    "Calcário": "suplemento",
    "Psyllium": "suplemento",
    "Dextrose": "suplemento",
    "Enteral": "enteral",
    "AIG": "enteral",
    "BSF": "enteral",
    "Complet": "enteral",
    "Ração": "commercial",
    "Patê": "commercial",
    "Sachê": "commercial",
    "Sucedâneo": "commercial",
    "Mucilon": "commercial",
    "Dieta": "commercial",
    "Farinha": "natural",
    "Creme": "natural",
    "AN": "natural",
    "Gordura": "natural",
    "Soja": "natural",
    "Coco": "natural",
    "Maizena": "natural",
    "Extrato": "natural",
}


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def slugify(value: str) -> str:
    value = strip_accents(value).lower()
    value = re.sub(r"[^a-z0-9]+", "-", value).strip("-")
    return re.sub(r"-{2,}", "-", value)


def compact_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def safe_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, datetime):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped.startswith("#"):
            return None
        normalized = stripped.replace("%", "").replace(",", ".")
        if re.fullmatch(r"-?\d+(?:\.\d+)?", normalized):
            return float(normalized)
    return None


def cell_raw_value(value: Any) -> str | float | int | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value
    return compact_spaces(str(value))


def parse_target(value: Any) -> dict[str, Any]:
    if value is None:
        return {"kind": "empty", "raw": None}
    if isinstance(value, datetime):
        return {"kind": "date_like", "raw": value.date().isoformat()}
    if isinstance(value, (int, float)):
        return {"kind": "number", "raw": value, "value": float(value)}

    raw = compact_spaces(str(value))
    if not raw:
        return {"kind": "empty", "raw": None}
    if raw.startswith("#"):
        return {"kind": "error", "raw": raw}

    normalized = raw.replace(",", ".")
    range_match = re.search(
        r"(?P<min>-?\d+(?:\.\d+)?)\s*(?:a|até|-)\s*(?P<max>-?\d+(?:\.\d+)?)",
        normalized,
        re.IGNORECASE,
    )
    if range_match:
        return {
            "kind": "range",
            "raw": raw,
            "min": float(range_match.group("min")),
            "max": float(range_match.group("max")),
        }

    comparator_match = re.search(r"^(?P<op>>=|<=|>|<|≥|≤)\s*(?P<value>-?\d+(?:\.\d+)?)", normalized)
    if comparator_match:
        op = comparator_match.group("op").replace("≥", ">=").replace("≤", "<=")
        return {
            "kind": "comparator",
            "raw": raw,
            "operator": op,
            "value": float(comparator_match.group("value")),
        }

    numeric_match = re.fullmatch(r"-?\d+(?:\.\d+)?", normalized)
    if numeric_match:
        return {"kind": "number", "raw": raw, "value": float(normalized)}

    unit_numeric_match = re.match(r"^(?P<num>-?\d+(?:\.\d+)?)\s*[A-Za-zµ/%]", normalized)
    if unit_numeric_match:
        return {
            "kind": "number_with_text",
            "raw": raw,
            "value": float(unit_numeric_match.group("num")),
        }

    return {"kind": "text", "raw": raw}


def infer_species(*values: str | None) -> str:
    haystack = strip_accents(" ".join(v for v in values if v)).lower()
    dog = any(token in haystack for token in ["cao", "caes", "cachorro", "canis", "canino"])
    cat = any(token in haystack for token in ["gato", "gatos", "felino", "felinos"])
    if dog and cat:
        return "both"
    if dog:
        return "dog"
    if cat:
        return "cat"
    return "unknown"


def infer_food_type(category: str | None, name: str) -> str:
    normalized = LOWERCASE_CATEGORY_FIXES.get(compact_spaces(category or ""), compact_spaces(category or ""))
    if normalized in MANUAL_FOOD_TYPE_RULES:
        return MANUAL_FOOD_TYPE_RULES[normalized]
    lowered = strip_accents(name).lower()
    if "enteral" in lowered:
        return "enteral"
    if "suplement" in lowered or "nutra" in lowered:
        return "suplemento"
    if normalized:
        return "commercial"
    return "unknown"


def infer_presentation(category: str | None, name: str) -> str:
    normalized = LOWERCASE_CATEGORY_FIXES.get(compact_spaces(category or ""), compact_spaces(category or ""))
    if normalized:
        return normalized
    lowered = strip_accents(name).lower()
    if "pate" in lowered:
        return "Patê"
    if "sache" in lowered:
        return "Sachê"
    return "Desconhecido"


def infer_basis_type(profile_label: str) -> str:
    normalized = strip_accents(profile_label).lower()
    if "%contribuicao energetica" in normalized:
        return "energy_percent"
    if "%ms" in normalized or "% ms" in normalized or normalized.endswith("ms%"):
        return "percent_dm"
    if "g/pv^" in normalized:
        return "per_metabolic_bw"
    if "g/kgpv" in normalized or "g/kgpv" in normalized.replace(" ", ""):
        return "per_kg_bw"
    if "100 kcal" in normalized:
        return "per_100kcal"
    return "range_text"


def infer_condition(profile_label: str) -> str | None:
    cleaned = re.sub(
        r"\s*-\s*(%MS|MS%|g/PV\^0,75|g/PV\^0,67|%Contribuição Energética|100 kcal)\s*$",
        "",
        profile_label,
        flags=re.IGNORECASE,
    )
    return compact_spaces(cleaned) or None


def infer_life_stage(profile_label: str) -> str | None:
    normalized = strip_accents(profile_label).lower()
    if "filhote" in normalized:
        return "growth"
    if "adult" in normalized:
        return "adult"
    if "gesta" in normalized:
        return "gestation"
    return None


def build_nutrient_catalog() -> list[dict[str, Any]]:
    catalog: list[dict[str, Any]] = []
    seen: set[str] = set()
    for header_map in (FOOD_HEADER_MAP, REQUIREMENT_HEADER_MAP):
        for label, key in header_map.items():
            if key in {"name", "category", "sourceLabel", "profileLabel", "vitaminSection", "extraRecommendations"}:
                continue
            if key in seen:
                continue
            seen.add(key)
            catalog.append(
                {
                    "key": key,
                    "label": label,
                    "unit": NUTRIENT_UNITS.get(key),
                }
            )
    return catalog


@dataclass
class WorkbookIssue:
    severity: str
    sheet: str
    message: str
    cell: str | None = None

    def as_dict(self) -> dict[str, Any]:
        payload = {
            "severity": self.severity,
            "sheet": self.sheet,
            "message": self.message,
        }
        if self.cell:
            payload["cell"] = self.cell
        return payload


def find_workbook(path_arg: str | None) -> Path:
    if path_arg:
        path = Path(path_arg)
        if path.exists():
            return path
        raise FileNotFoundError(f"Workbook not found: {path}")

    downloads = Path.home() / "Downloads"
    matches = sorted(downloads.glob(DEFAULT_GLOB))
    if not matches:
        raise FileNotFoundError(f"No workbook matching {DEFAULT_GLOB} in {downloads}")
    return matches[0]


def build_foods(wb_formula: Any, wb_values: Any, issues: list[WorkbookIssue]) -> list[dict[str, Any]]:
    mn_formula = wb_formula["Alimentos MN"]
    mn_values = wb_values["Alimentos MN"]
    ms_values = wb_values["Alimentos MS"]

    mn_rows: list[dict[str, Any]] = []
    ms_lookup: dict[str, tuple[int, Any, Any]] = {}

    for row in range(2, ms_values.max_row + 1):
        name = ms_values.cell(row, 1).value
        if name:
            ms_lookup[compact_spaces(str(name))] = (row, ms_values, wb_formula["Alimentos MS"])

    category_counter: Counter[str] = Counter()

    for row in range(2, mn_values.max_row + 1):
        name_value = mn_values.cell(row, 1).value
        if not name_value:
            continue

        name = compact_spaces(str(name_value))
        ms_row_tuple = ms_lookup.get(name)
        raw_category = cell_raw_value(mn_values.cell(row, 2).value)
        category = compact_spaces(str(raw_category)) if raw_category else None
        if category:
            category_counter[category] += 1

        normalized_category = LOWERCASE_CATEGORY_FIXES.get(category or "", category)
        if category != normalized_category and category:
            issues.append(
                WorkbookIssue(
                    severity="warning",
                    sheet="Alimentos MN",
                    cell=f"B{row}",
                    message=f"Categoria inconsistente '{category}' normalizada para '{normalized_category}'.",
                )
            )
        if category is None:
            issues.append(
                WorkbookIssue(
                    severity="warning",
                    sheet="Alimentos MN",
                    cell=f"B{row}",
                    message=f"Alimento '{name}' sem categoria preenchida.",
                )
            )

        nutrients_as_fed: dict[str, float | None] = {}
        nutrients_dry_matter: dict[str, float | None] = {}
        missing_keys: list[str] = []
        dry_matter_pct = safe_number(mn_values.cell(row, 4).value)

        for col, header in enumerate(
            [mn_values.cell(1, c).value for c in range(3, mn_values.max_column + 1)],
            start=3,
        ):
            if header not in FOOD_HEADER_MAP:
                continue
            key = FOOD_HEADER_MAP[header]
            as_fed_value = safe_number(mn_values.cell(row, col).value)
            nutrients_as_fed[key] = as_fed_value

            if as_fed_value is None:
                nutrients_dry_matter[key] = None
                if key not in {"moisturePct", "dryMatterPct"}:
                    missing_keys.append(key)
                continue

            if key == "moisturePct":
                nutrients_dry_matter[key] = 0.0 if dry_matter_pct else None
                continue
            if key == "dryMatterPct":
                nutrients_dry_matter[key] = 100.0 if dry_matter_pct else None
                continue

            if dry_matter_pct in (None, 0):
                nutrients_dry_matter[key] = None
                if as_fed_value:
                    issues.append(
                        WorkbookIssue(
                            severity="warning",
                            sheet="Alimentos MN",
                            cell=f"{get_column_letter(col)}{row}",
                            message=f"Matéria seca zerada ou ausente para '{name}', conversão MS não aplicada.",
                        )
                    )
            else:
                nutrients_dry_matter[key] = round((as_fed_value * 100) / dry_matter_pct, 6)

        if name in ms_lookup:
            ms_row = ms_row_tuple[0]
            for check_col in range(3, min(ms_values.max_column, 40) + 1):
                value = ms_values.cell(ms_row, check_col).value
                if isinstance(value, str) and value.startswith("#"):
                    issues.append(
                        WorkbookIssue(
                            severity="warning",
                            sheet="Alimentos MS",
                            cell=f"{get_column_letter(check_col)}{ms_row}",
                            message=f"Erro em célula calculada de matéria seca para '{name}': {value}. A importação usa conversão derivada do MN.",
                        )
                    )

        mn_rows.append(
            {
                "id": slugify(name),
                "slug": slugify(name),
                "name": name,
                "category": category,
                "categoryNormalized": normalized_category,
                "sourceSheet": "Alimentos MN",
                "sourceReference": {
                    "workbook": "Planilha de Formulação e Avaliação de Dietas - GENUTRI.xlsx",
                    "mnRow": row,
                    "msRow": ms_row_tuple[0] if ms_row_tuple else None,
                },
                "speciesScope": infer_species(name, category),
                "foodType": infer_food_type(normalized_category, name),
                "presentation": infer_presentation(normalized_category, name),
                "nutrientsAsFed": nutrients_as_fed,
                "nutrientsDryMatter": nutrients_dry_matter,
                "missingNutrients": sorted(set(missing_keys)),
                "notes": [],
            }
        )

    if category_counter.get("None"):
        issues.append(
            WorkbookIssue(
                severity="warning",
                sheet="Alimentos MN",
                message="Existe ao menos um alimento sem categoria, exigindo revisão manual.",
            )
        )

    return mn_rows


def build_requirement_profiles(wb_values: Any, issues: list[WorkbookIssue]) -> list[dict[str, Any]]:
    title = next(sheet for sheet in wb_values.sheetnames if "Exig" in sheet and "Ener" not in sheet)
    ws = wb_values[title]
    profiles: list[dict[str, Any]] = []

    for row in range(2, ws.max_row + 1):
        source = cell_raw_value(ws.cell(row, 1).value)
        label = cell_raw_value(ws.cell(row, 2).value)
        if not source and not label:
            continue

        label_text = str(label or "").strip()
        nutrient_targets: dict[str, dict[str, Any]] = {}

        for col in range(3, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header is None:
                continue
            key = REQUIREMENT_HEADER_MAP.get(header)
            if not key:
                continue
            parsed = parse_target(ws.cell(row, col).value)
            nutrient_targets[key] = parsed
            if parsed["kind"] == "date_like":
                issues.append(
                    WorkbookIssue(
                        severity="warning",
                        sheet=title,
                        cell=f"{get_column_letter(col)}{row}",
                        message=f"Campo '{header}' da exigência '{label_text}' veio como data ({parsed['raw']}) e será tratado como comparação assistida/manual.",
                    )
                )

        extras = {
            "urinaryPhCat": nutrient_targets.get("urinaryPhCat"),
            "urinaryDensity": nutrient_targets.get("urinaryDensity"),
            "extraRecommendations": nutrient_targets.get("extraRecommendations"),
        }

        profiles.append(
            {
                "id": slugify(f"{source}-{label}"),
                "source": source,
                "label": label,
                "species": infer_species(str(source or ""), label_text),
                "lifeStage": infer_life_stage(label_text),
                "condition": infer_condition(label_text),
                "basisType": infer_basis_type(label_text),
                "nutrientTargets": nutrient_targets,
                "extras": extras,
                "sourceReference": {
                    "sheet": title,
                    "row": row,
                },
            }
        )

    return profiles


def build_energy_rules(wb_values: Any) -> list[dict[str, Any]]:
    title = next(sheet for sheet in wb_values.sheetnames if "Ener" in sheet)
    ws = wb_values[title]
    rules: list[dict[str, Any]] = []

    for row in range(7, ws.max_row + 1):
        label = ws.cell(row, 1).value
        exponent = safe_number(ws.cell(row, 2).value)
        constant = safe_number(ws.cell(row, 3).value)
        if not label or exponent is None or constant is None:
            continue
        species = infer_species(str(label))
        rules.append(
            {
                "id": slugify(str(label)),
                "species": "cat" if "gato" in strip_accents(str(label)).lower() else "dog",
                "sourceLabel": compact_spaces(str(label)),
                "exponent": exponent,
                "constant": constant,
                "formulaLabel": f"{constant:g} * peso^{exponent:g}",
                "source": "GENUTRI / Exigências Energética",
                "sourceReference": {
                    "sheet": title,
                    "row": row,
                },
            }
        )

    return rules


def build_sheet_summary(wb_formula: Any, wb_values: Any) -> list[dict[str, Any]]:
    summary: list[dict[str, Any]] = []
    for ws in wb_formula.worksheets:
        data_ws = wb_values[ws.title]
        formula_count = 0
        error_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
        for row in data_ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#"):
                    error_count += 1
        summary.append(
            {
                "title": ws.title,
                "dimension": ws.calculate_dimension(),
                "formulaCells": formula_count,
                "mergedRanges": len(ws.merged_cells.ranges),
                "errorValues": error_count,
            }
        )
    return summary


def build_audit_markdown(dataset: dict[str, Any]) -> str:
    lines: list[str] = []
    lines.append("# Auditoria GENUTRI para Energia Vet")
    lines.append("")
    lines.append(f"- Gerado em: `{dataset['meta']['generatedAt']}`")
    lines.append(f"- Workbook: `{dataset['meta']['sourceWorkbook']}`")
    lines.append(f"- Alimentos importados: `{len(dataset['foods'])}`")
    lines.append(f"- Perfis de exigência importados: `{len(dataset['requirements'])}`")
    lines.append(f"- Regras energéticas importadas: `{len(dataset['energyRules'])}`")
    lines.append("")
    lines.append("## Abas")
    lines.append("")
    for sheet in dataset["audit"]["sheetSummary"]:
        lines.append(
            f"- `{sheet['title']}`: dimensão `{sheet['dimension']}`, fórmulas `{sheet['formulaCells']}`, erros calculados `{sheet['errorValues']}`."
        )
    lines.append("")
    lines.append("## Problemas Detectados")
    lines.append("")
    if dataset["audit"]["issues"]:
        for issue in dataset["audit"]["issues"]:
            cell = f" `{issue['cell']}`" if issue.get("cell") else ""
            lines.append(f"- [{issue['severity'].upper()}] `{issue['sheet']}`{cell}: {issue['message']}")
    else:
        lines.append("- Nenhum problema detectado.")
    return "\n".join(lines) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa a planilha GENUTRI para o dataset do Energia Vet.")
    parser.add_argument("--input", dest="input_path", help="Caminho do arquivo .xlsx")
    parser.add_argument("--output", dest="output_path", help="Caminho do JSON gerado")
    parser.add_argument("--audit-output", dest="audit_output_path", help="Caminho do relatório markdown")
    args = parser.parse_args()

    workbook_path = find_workbook(args.input_path)
    output_path = Path(args.output_path) if args.output_path else DEFAULT_OUTPUT
    audit_path = Path(args.audit_output_path) if args.audit_output_path else DEFAULT_AUDIT

    wb_formula = load_workbook(workbook_path, data_only=False)
    wb_values = load_workbook(workbook_path, data_only=True)
    issues: list[WorkbookIssue] = []

    foods = build_foods(wb_formula, wb_values, issues)
    requirements = build_requirement_profiles(wb_values, issues)
    energy_rules = build_energy_rules(wb_values)

    issues.append(
        WorkbookIssue(
            severity="warning",
            sheet="Alimentos MS",
            message="A aba Alimentos MS usa IFERROR e materializa zeros onde o MN está vazio. O importador recalcula matéria seca preservando null para não inventar nutrientes.",
        )
    )
    issues.append(
        WorkbookIssue(
            severity="warning",
            sheet="Ficha Nutricional",
            cell="B24",
            message="A célula B24 contém 'Z', causando #VALUE! em B25:B31. O app gera plano alimentar sem replicar esse bug.",
        )
    )
    issues.append(
        WorkbookIssue(
            severity="info",
            sheet="Prompt",
            message="O prompt referencia um PDF FEDIAF, mas o arquivo não foi encontrado automaticamente no diretório Downloads durante esta importação.",
        )
    )

    dataset = {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "sourceWorkbook": str(workbook_path),
            "nutrientCatalog": build_nutrient_catalog(),
        },
        "foods": foods,
        "requirements": requirements,
        "energyRules": energy_rules,
        "audit": {
            "sheetSummary": build_sheet_summary(wb_formula, wb_values),
            "issues": [issue.as_dict() for issue in issues],
        },
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(dataset, ensure_ascii=False, indent=2), encoding="utf-8")

    audit_path.parent.mkdir(parents=True, exist_ok=True)
    audit_path.write_text(build_audit_markdown(dataset), encoding="utf-8")

    print(f"Workbook: {workbook_path}")
    print(f"Foods: {len(foods)}")
    print(f"Requirements: {len(requirements)}")
    print(f"Energy rules: {len(energy_rules)}")
    print(f"Dataset written to: {output_path}")
    print(f"Audit written to: {audit_path}")


if __name__ == "__main__":
    main()
